from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np


ANGLE_HEADER_KEYWORDS = ("theta", "angle", "azimuth", "az", "deg")
GAIN_HEADER_KEYWORDS = ("gain", "db", "dbi", "realized")
PATTERN_KIND_AMPLITUDE = "amplitude"
PATTERN_KIND_PHASE = "phase"
PATTERN_PLANE_HORIZONTAL = "horizontal"
PATTERN_PLANE_ELEVATION = "elevation"
PHASE_CALIBRATION_ANGLE_DEG = 0.0


@dataclass(frozen=True)
class ElementPattern:
    name: str
    source_path: str
    angle_column: str
    horizontal_column: str
    elevation_column: str | None
    angles_deg: np.ndarray
    horizontal_gain_db: np.ndarray
    elevation_gain_db: np.ndarray | None = None

    def horizontal_gain_db_at(self, angles_deg: np.ndarray) -> np.ndarray:
        return np.interp(
            angles_deg,
            self.angles_deg,
            self.horizontal_gain_db,
            left=float(self.horizontal_gain_db[0]),
            right=float(self.horizontal_gain_db[-1]),
        )

    def elevation_gain_db_at(self, angles_deg: np.ndarray) -> np.ndarray:
        gain_db = (
            self.elevation_gain_db
            if self.elevation_gain_db is not None
            else self.horizontal_gain_db
        )
        return np.interp(
            angles_deg,
            self.angles_deg,
            gain_db,
            left=float(gain_db[0]),
            right=float(gain_db[-1]),
        )

    def normalized_horizontal_gain_db_at(self, angles_deg: np.ndarray) -> np.ndarray:
        return self.horizontal_gain_db_at(angles_deg) - float(np.max(self.horizontal_gain_db))

    def normalized_elevation_gain_db_at(self, angles_deg: np.ndarray) -> np.ndarray:
        gain_db = (
            self.elevation_gain_db
            if self.elevation_gain_db is not None
            else self.horizontal_gain_db
        )
        return self.elevation_gain_db_at(angles_deg) - float(np.max(gain_db))

    def amplitude_grid(
        self,
        azimuth_deg: np.ndarray,
        elevation_deg: np.ndarray,
    ) -> np.ndarray:
        horizontal = np.power(10.0, self.horizontal_gain_db_at(azimuth_deg) / 20.0)
        if self.elevation_gain_db is None:
            return horizontal
        elevation = np.power(10.0, self.elevation_gain_db_at(elevation_deg) / 20.0)
        return horizontal * elevation

    def display_name(self) -> str:
        if self.elevation_column is None:
            return f"{self.name} ({self.horizontal_column})"
        return f"{self.name} (H: {self.horizontal_column}; V: {self.elevation_column})"

    def swapped_axes(self) -> "ElementPattern":
        if self.elevation_column is None or self.elevation_gain_db is None:
            return self
        return ElementPattern(
            name=self.name,
            source_path=self.source_path,
            angle_column=self.angle_column,
            horizontal_column=self.elevation_column,
            elevation_column=self.horizontal_column,
            angles_deg=self.angles_deg,
            horizontal_gain_db=self.elevation_gain_db,
            elevation_gain_db=self.horizontal_gain_db,
        )


@dataclass(frozen=True)
class PatternSeries:
    name: str
    source_path: str
    angle_column: str
    value_column: str
    value_kind: str
    angles_deg: np.ndarray
    values: np.ndarray

    def values_at(self, angles_deg: np.ndarray) -> np.ndarray:
        query_angles = np.asarray(angles_deg, dtype=float)
        values = self.values
        if self.value_kind == PATTERN_KIND_PHASE:
            unwrapped_rad = np.unwrap(np.radians(values))
            interpolated_rad = np.interp(
                query_angles,
                self.angles_deg,
                unwrapped_rad,
                left=float(unwrapped_rad[0]),
                right=float(unwrapped_rad[-1]),
            )
            return np.degrees(interpolated_rad)
        return np.interp(
            query_angles,
            self.angles_deg,
            values,
            left=float(values[0]),
            right=float(values[-1]),
        )

    def short_label(self) -> str:
        return f"{Path(self.source_path).name}:{self.value_column}"


@dataclass(frozen=True)
class ChannelPattern:
    amplitude_horizontal: PatternSeries | None = None
    amplitude_elevation: PatternSeries | None = None
    phase_horizontal: PatternSeries | None = None
    phase_elevation: PatternSeries | None = None

    def is_empty(self) -> bool:
        return all(series is None for series in self._all_series())

    def series_count(self) -> int:
        return sum(series is not None for series in self._all_series())

    def with_series(
        self,
        kind: str,
        plane: str,
        series: PatternSeries | None,
    ) -> "ChannelPattern":
        values = {
            "amplitude_horizontal": self.amplitude_horizontal,
            "amplitude_elevation": self.amplitude_elevation,
            "phase_horizontal": self.phase_horizontal,
            "phase_elevation": self.phase_elevation,
        }
        values[_channel_pattern_field(kind, plane)] = series
        return ChannelPattern(**values)

    def series_for(self, kind: str, plane: str) -> PatternSeries | None:
        return getattr(self, _channel_pattern_field(kind, plane))

    def _all_series(self) -> tuple[PatternSeries | None, ...]:
        return (
            self.amplitude_horizontal,
            self.amplitude_elevation,
            self.phase_horizontal,
            self.phase_elevation,
        )


@dataclass
class ChannelPatternSet:
    patterns: dict[str, ChannelPattern]

    def __init__(
        self, patterns: dict[str, ChannelPattern] | None = None
    ) -> None:
        self.patterns = dict(patterns or {})

    def is_empty(self) -> bool:
        return not any(not pattern.is_empty() for pattern in self.patterns.values())

    def set_series(
        self,
        channel_name: str,
        kind: str,
        plane: str,
        series: PatternSeries | None,
    ) -> None:
        current = self.patterns.get(channel_name, ChannelPattern())
        updated = current.with_series(kind, plane, series)
        if updated.is_empty():
            self.patterns.pop(channel_name, None)
        else:
            self.patterns[channel_name] = updated

    def update_many(
        self,
        series_by_channel: dict[str, PatternSeries],
        kind: str,
        plane: str,
    ) -> None:
        for channel_name, series in series_by_channel.items():
            self.set_series(channel_name, kind, plane, series)

    def clear_channel(self, channel_name: str) -> None:
        self.patterns.pop(channel_name, None)

    def clear(self) -> None:
        self.patterns.clear()

    def pattern_for(self, channel_name: str) -> ChannelPattern:
        return self.patterns.get(channel_name, ChannelPattern())

    def configured_channel_count(self) -> int:
        return sum(not pattern.is_empty() for pattern in self.patterns.values())

    def configured_series_count(self) -> int:
        return sum(pattern.series_count() for pattern in self.patterns.values())

    def complex_weights(
        self,
        channel_names: list[str],
        azimuth_deg: np.ndarray,
        elevation_deg: np.ndarray,
    ) -> np.ndarray:
        azimuth_grid, elevation_grid = np.broadcast_arrays(
            np.asarray(azimuth_deg, dtype=float),
            np.asarray(elevation_deg, dtype=float),
        )
        weights = np.ones((len(channel_names), *azimuth_grid.shape), dtype=complex)
        if self.is_empty():
            return weights

        for channel_index, channel_name in enumerate(channel_names):
            pattern = self.pattern_for(channel_name)
            amplitude_db = np.zeros_like(azimuth_grid, dtype=float)
            phase_deg = np.zeros_like(azimuth_grid, dtype=float)

            if pattern.amplitude_horizontal is not None:
                amplitude_db += pattern.amplitude_horizontal.values_at(azimuth_grid)
            if pattern.amplitude_elevation is not None:
                amplitude_db += pattern.amplitude_elevation.values_at(elevation_grid)
            if pattern.phase_horizontal is not None:
                phase_deg += pattern.phase_horizontal.values_at(azimuth_grid)
            if pattern.phase_elevation is not None:
                phase_deg += pattern.phase_elevation.values_at(elevation_grid)

            weights[channel_index] = np.power(10.0, amplitude_db / 20.0) * np.exp(
                1j * np.radians(phase_deg)
            )

        return weights


@dataclass(frozen=True)
class PatternCutMetrics:
    peak_angle_deg: float
    peak_gain_dbi: float
    beamwidth_3db_deg: float | None
    beamwidth_6db_deg: float | None


def load_element_pattern(path: str | Path) -> ElementPattern:
    source_path = Path(path)
    rows = _read_table_rows(source_path)

    if len(rows) < 2:
        raise ValueError("Pattern file must contain a header and at least one data row.")

    header = [cell.strip() for cell in rows[0]]
    angle_index = _find_angle_column(header)
    gain_indices = _find_gain_columns(header, exclude={angle_index})
    if not gain_indices:
        raise ValueError("Pattern file must contain at least one gain column.")
    horizontal_index, elevation_index = _default_pattern_axes(gain_indices)

    values = []
    for row_index, row in enumerate(rows[1:], start=2):
        required_columns = [angle_index, horizontal_index]
        if elevation_index is not None:
            required_columns.append(elevation_index)
        required_index = max(required_columns)
        if required_index >= len(row):
            continue
        try:
            angle = _parse_float(row[angle_index])
            horizontal_gain = _parse_float(row[horizontal_index])
            elevation_gain = (
                _parse_float(row[elevation_index])
                if elevation_index is not None
                else np.nan
            )
        except ValueError as exc:
            raise ValueError(f"Invalid numeric value on row {row_index}.") from exc
        if np.isfinite(angle) and np.isfinite(horizontal_gain):
            values.append((angle, horizontal_gain, elevation_gain))

    if len(values) < 2:
        raise ValueError("Pattern file must contain at least two valid data rows.")

    values_array = np.asarray(values, dtype=float)
    order = np.argsort(values_array[:, 0])
    angles = values_array[order, 0]
    horizontal_gains = values_array[order, 1]
    elevation_gains = values_array[order, 2]
    angles, horizontal_gains, elevation_gains = _merge_duplicate_angles(
        angles, horizontal_gains, elevation_gains
    )
    elevation_gain_db = elevation_gains if elevation_index is not None else None

    return ElementPattern(
        name=source_path.name,
        source_path=str(source_path),
        angle_column=header[angle_index],
        horizontal_column=header[horizontal_index],
        elevation_column=header[elevation_index] if elevation_index is not None else None,
        angles_deg=angles,
        horizontal_gain_db=horizontal_gains,
        elevation_gain_db=elevation_gain_db,
    )


def load_hfss_pattern_series(
    path: str | Path,
    value_kind: str,
    column_offset: int = 0,
) -> PatternSeries:
    series = load_hfss_pattern_series_columns(
        path,
        value_kind=value_kind,
        max_columns=column_offset + 1,
    )
    if column_offset >= len(series):
        raise ValueError(
            f"Pattern file has {len(series)} data column(s) after Theta; "
            f"column {column_offset + 1} was requested."
        )
    return series[column_offset]


def load_hfss_summary_pattern(
    path: str | Path,
    channel_names: list[str],
    value_kind: str,
) -> dict[str, PatternSeries]:
    series = load_hfss_pattern_series_columns(
        path,
        value_kind=value_kind,
        max_columns=len(channel_names),
    )
    mapped_channel_names = _summary_pattern_channel_names(channel_names, len(series))
    if mapped_channel_names is None:
        raise ValueError(
            f"Summary pattern has {len(series)} data column(s) after Theta, "
            f"but the current layout needs {len(channel_names)} physical channels."
        )
    return {
        channel_name: pattern_series
        for channel_name, pattern_series in zip(mapped_channel_names, series)
    }


def _summary_pattern_channel_names(
    channel_names: list[str],
    data_column_count: int,
) -> list[str] | None:
    if data_column_count == len(channel_names):
        return list(channel_names)

    tx_names = [
        channel_name
        for channel_name in channel_names
        if channel_name.strip().lower().startswith("tx")
    ]
    rx_names = [
        channel_name
        for channel_name in channel_names
        if channel_name.strip().lower().startswith("rx")
    ]
    if rx_names and data_column_count == len(rx_names):
        return rx_names
    if tx_names and rx_names and data_column_count == len(rx_names) + 1:
        return [tx_names[0], *rx_names]
    return None


def load_hfss_pattern_series_columns(
    path: str | Path,
    value_kind: str,
    max_columns: int | None = None,
) -> list[PatternSeries]:
    if value_kind not in {PATTERN_KIND_AMPLITUDE, PATTERN_KIND_PHASE}:
        raise ValueError(f"Unknown pattern value kind: {value_kind!r}")

    source_path = Path(path)
    rows = _read_table_rows(source_path)

    if len(rows) < 2:
        raise ValueError("Pattern file must contain a header and at least one data row.")

    header = [cell.strip() for cell in rows[0]]
    angle_index = _find_angle_column(header)
    data_indices = _hfss_data_columns_after_angle(header, angle_index)
    if max_columns is not None:
        data_indices = data_indices[:max_columns]
    if not data_indices:
        raise ValueError("Pattern file must contain at least one data column after Theta.")

    values = []
    for row_index, row in enumerate(rows[1:], start=2):
        required_index = max([angle_index, *data_indices])
        if required_index >= len(row):
            continue
        try:
            angle = _parse_float(row[angle_index])
            data_values = [
                _parse_hfss_pattern_value(row[index], value_kind)
                for index in data_indices
            ]
        except ValueError as exc:
            raise ValueError(f"Invalid numeric value on row {row_index}.") from exc
        if np.isfinite(angle) and all(np.isfinite(value) for value in data_values):
            values.append((angle, *data_values))

    if len(values) < 2:
        raise ValueError("Pattern file must contain at least two valid data rows.")

    values_array = np.asarray(values, dtype=float)
    order = np.argsort(values_array[:, 0])
    angles = values_array[order, 0]
    data_matrix = values_array[order, 1:]
    angles, data_matrix = _merge_duplicate_series_values(angles, data_matrix)
    if value_kind == PATTERN_KIND_PHASE:
        data_matrix = _calibrate_phase_matrix(
            angles,
            data_matrix,
            reference_angle_deg=PHASE_CALIBRATION_ANGLE_DEG,
        )

    return [
        PatternSeries(
            name=source_path.name,
            source_path=str(source_path),
            angle_column=header[angle_index],
            value_column=header[column_index],
            value_kind=value_kind,
            angles_deg=angles,
            values=data_matrix[:, data_index],
        )
        for data_index, column_index in enumerate(data_indices)
    ]


def available_gain_columns(path: str | Path) -> list[str]:
    source_path = Path(path)
    rows = _read_table_rows(source_path)
    if not rows:
        return []
    header = [cell.strip() for cell in rows[0]]
    angle_index = _find_angle_column(header)
    return [header[index] for index in _find_gain_columns(header, exclude={angle_index})]


def pattern_cut_metrics(
    angles_deg: np.ndarray,
    gain_db: np.ndarray,
) -> PatternCutMetrics:
    if len(angles_deg) == 0 or len(gain_db) == 0:
        raise ValueError("Pattern cut must contain at least one point.")

    peak_index = int(np.argmax(gain_db))
    peak_angle = float(angles_deg[peak_index])
    peak_gain = float(gain_db[peak_index])
    return PatternCutMetrics(
        peak_angle_deg=peak_angle,
        peak_gain_dbi=peak_gain,
        beamwidth_3db_deg=_beamwidth_at_drop(angles_deg, gain_db, peak_index, 3.0),
        beamwidth_6db_deg=_beamwidth_at_drop(angles_deg, gain_db, peak_index, 6.0),
    )


def format_pattern_cut_metrics(metrics: PatternCutMetrics) -> str:
    bw_3db = _format_optional_degrees(metrics.beamwidth_3db_deg)
    bw_6db = _format_optional_degrees(metrics.beamwidth_6db_deg)
    return (
        f"Peak {metrics.peak_gain_dbi:.2f} dBi @ {metrics.peak_angle_deg:.1f}° | "
        f"3dB BW {bw_3db} | 6dB BW {bw_6db}"
    )


def _detect_delimiter(path: Path) -> str:
    if path.suffix.lower() == ".tsv":
        return "\t"

    sample = path.read_text(encoding="utf-8-sig", errors="ignore")[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        return ","
    return dialect.delimiter


def _read_table_rows(path: Path) -> list[list[str]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Reading XLSX pattern files requires openpyxl.") from exc
        workbook = load_workbook(path, data_only=True, read_only=True)
        sheet = workbook.active
        return [
            ["" if cell is None else str(cell).strip() for cell in row]
            for row in sheet.iter_rows(values_only=True)
            if row and not _is_comment_or_blank(["" if cell is None else str(cell) for cell in row])
        ]

    delimiter = _detect_delimiter(path)
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return [
            row
            for row in csv.reader(file, delimiter=delimiter)
            if row and not _is_comment_or_blank(row)
        ]


def _beamwidth_at_drop(
    angles_deg: np.ndarray,
    gain_db: np.ndarray,
    peak_index: int,
    drop_db: float,
) -> float | None:
    threshold = float(gain_db[peak_index]) - drop_db
    left = _threshold_angle(angles_deg, gain_db, peak_index, threshold, -1)
    right = _threshold_angle(angles_deg, gain_db, peak_index, threshold, 1)
    if left is None or right is None:
        return None
    return float(right - left)


def _threshold_angle(
    angles_deg: np.ndarray,
    gain_db: np.ndarray,
    peak_index: int,
    threshold_db: float,
    direction: int,
) -> float | None:
    index = peak_index + direction
    while 0 <= index < len(gain_db):
        if gain_db[index] <= threshold_db:
            prev_index = index - direction
            x0 = float(angles_deg[prev_index])
            x1 = float(angles_deg[index])
            y0 = float(gain_db[prev_index])
            y1 = float(gain_db[index])
            if y1 == y0:
                return x1
            fraction = (threshold_db - y0) / (y1 - y0)
            return x0 + fraction * (x1 - x0)
        index += direction
    return None


def _format_optional_degrees(value: float | None) -> str:
    if value is None or not np.isfinite(value):
        return "N/A"
    return f"{value:.1f}°"


def _find_angle_column(header: list[str]) -> int:
    for index, column in enumerate(header):
        text = column.lower()
        if "theta" in text:
            return index
    for index, column in enumerate(header):
        text = column.lower()
        if any(keyword in text for keyword in ANGLE_HEADER_KEYWORDS):
            return index
    return 0


def _find_gain_columns(header: list[str], exclude: set[int]) -> list[int]:
    preferred = [
        index
        for index, column in enumerate(header)
        if index not in exclude and _looks_like_gain_column(column)
    ]
    if preferred:
        return preferred
    return [index for index in range(len(header)) if index not in exclude]


def _hfss_data_columns_after_angle(header: list[str], angle_index: int) -> list[int]:
    after_angle = list(range(angle_index + 1, len(header)))
    if after_angle:
        return after_angle
    return [
        index
        for index, column in enumerate(header)
        if index != angle_index
        and "freq" not in column.lower()
        and "phi" not in column.lower()
    ]


def _channel_pattern_field(kind: str, plane: str) -> str:
    if kind == PATTERN_KIND_AMPLITUDE and plane == PATTERN_PLANE_HORIZONTAL:
        return "amplitude_horizontal"
    if kind == PATTERN_KIND_AMPLITUDE and plane == PATTERN_PLANE_ELEVATION:
        return "amplitude_elevation"
    if kind == PATTERN_KIND_PHASE and plane == PATTERN_PLANE_HORIZONTAL:
        return "phase_horizontal"
    if kind == PATTERN_KIND_PHASE and plane == PATTERN_PLANE_ELEVATION:
        return "phase_elevation"
    raise ValueError(f"Unknown channel pattern slot: kind={kind!r}, plane={plane!r}")


def _default_pattern_axes(gain_indices: list[int]) -> tuple[int, int | None]:
    if len(gain_indices) < 2:
        return gain_indices[0], None
    return gain_indices[1], gain_indices[0]


def _looks_like_gain_column(column: str) -> bool:
    text = column.lower()
    return any(keyword in text for keyword in GAIN_HEADER_KEYWORDS)


def _parse_float(value: str) -> float:
    cleaned = value.strip()
    if not cleaned:
        raise ValueError("empty numeric value")
    return float(cleaned)


def _parse_hfss_pattern_value(value: str, value_kind: str) -> float:
    if value_kind == PATTERN_KIND_PHASE and _looks_like_complex(value):
        return _complex_phase_degrees(value)
    return _parse_float(value)


def _complex_phase_degrees(value: str) -> float:
    return float(np.degrees(np.angle(_parse_complex_value(value))))


def _parse_complex_value(value: str) -> complex:
    text = value.strip()
    if not text:
        raise ValueError("empty complex value")
    normalized = (
        text.replace("i", "j")
        .replace("I", "j")
        .replace("−", "-")
        .replace("，", ",")
        .strip()
    )
    normalized = normalized.replace(" ", "")
    if "∠" in normalized:
        magnitude_text, phase_text = normalized.split("∠", 1)
        return float(magnitude_text) * np.exp(1j * np.radians(float(phase_text)))
    return complex(normalized)


def _looks_like_complex(value: str) -> bool:
    text = value.strip().lower()
    return "j" in text or "i" in text or "∠" in text


def _is_comment_or_blank(row: Iterable[str]) -> bool:
    stripped = [cell.strip() for cell in row]
    return not any(stripped) or stripped[0].startswith("#")


def _merge_duplicate_angles(
    angles: np.ndarray,
    horizontal_gains: np.ndarray,
    elevation_gains: np.ndarray,
) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    unique_angles, inverse = np.unique(angles, return_inverse=True)
    if len(unique_angles) == len(angles):
        return angles, horizontal_gains, elevation_gains

    merged_horizontal = np.zeros_like(unique_angles, dtype=float)
    merged_elevation = np.zeros_like(unique_angles, dtype=float)
    for index in range(len(unique_angles)):
        merged_horizontal[index] = float(np.mean(horizontal_gains[inverse == index]))
        merged_elevation[index] = float(np.mean(elevation_gains[inverse == index]))
    return unique_angles, merged_horizontal, merged_elevation


def _merge_duplicate_series_values(
    angles: np.ndarray,
    data_matrix: np.ndarray,
) -> tuple[np.ndarray, np.ndarray]:
    unique_angles, inverse = np.unique(angles, return_inverse=True)
    if len(unique_angles) == len(angles):
        return angles, data_matrix

    merged = np.zeros((len(unique_angles), data_matrix.shape[1]), dtype=float)
    for index in range(len(unique_angles)):
        merged[index] = np.mean(data_matrix[inverse == index], axis=0)
    return unique_angles, merged


def _calibrate_phase_matrix(
    angles: np.ndarray,
    phase_matrix_deg: np.ndarray,
    reference_angle_deg: float = PHASE_CALIBRATION_ANGLE_DEG,
) -> np.ndarray:
    low = float(angles[0])
    high = float(angles[-1])
    if reference_angle_deg < low or reference_angle_deg > high:
        raise ValueError(
            f"Phase pattern covers {low:g}..{high:g} deg and cannot be "
            f"calibrated at {reference_angle_deg:g} deg."
        )

    unwrapped_rad = np.unwrap(np.radians(phase_matrix_deg), axis=0)
    reference_rad = np.array(
        [
            np.interp(reference_angle_deg, angles, unwrapped_rad[:, column])
            for column in range(unwrapped_rad.shape[1])
        ],
        dtype=float,
    )
    return np.degrees(unwrapped_rad - reference_rad[None, :])
