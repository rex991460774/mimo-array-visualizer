from __future__ import annotations

import builtins

import numpy as np
import pytest
from openpyxl import Workbook

from virtual_array.analysis import calculate_metrics_and_psf, dbf_azimuth_spectrum
from virtual_array.element_pattern import (
    PATTERN_KIND_PHASE,
    PATTERN_PLANE_HORIZONTAL,
    ChannelPatternSet,
    ElementPattern,
    PatternSeries,
    format_pattern_cut_metrics,
    load_hfss_pattern_series,
    load_hfss_summary_pattern,
    load_element_pattern,
    pattern_cut_metrics,
    virtual_channel_names,
)
from virtual_array.geometry import AntennaArray


def test_loads_hfss_realized_gain_eh_plane_csv(tmp_path) -> None:
    csv_path = tmp_path / "pattern.csv"
    csv_path.write_text(
        '"Theta [deg]","dB(RealizedGainTotal) [] - Freq=\'77GHz\' Phi=\'0deg\'",'
        '"dB(RealizedGainTotal)_1 [] - Freq=\'77GHz\' Phi=\'90deg\'"\n'
        "-1,-10,-20\n"
        "0,0,-5\n"
        "1,-10,-20\n",
        encoding="utf-8",
    )

    pattern = load_element_pattern(csv_path)

    assert pattern.angle_column == "Theta [deg]"
    assert "Phi='90deg'" in pattern.horizontal_column
    assert pattern.elevation_column is not None
    assert "Phi='0deg'" in pattern.elevation_column
    assert np.array_equal(pattern.angles_deg, np.array([-1.0, 0.0, 1.0]))
    assert np.array_equal(pattern.horizontal_gain_db, np.array([-20.0, -5.0, -20.0]))
    assert np.array_equal(pattern.elevation_gain_db, np.array([-10.0, 0.0, -10.0]))


def test_loads_hfss_summary_columns_in_tx_rx_order_with_zero_phase_calibration(
    tmp_path,
) -> None:
    csv_path = tmp_path / "phase-summary.csv"
    csv_path.write_text(
        '"Freq [GHz]","Phi [deg]","Theta [deg]","cang_deg(rETheta) [deg]",'
        '"cang_deg(rETheta)_1 [deg]","cang_deg(rETheta)_2 [deg]",'
        '"cang_deg(rETheta)_3 [deg]"\n'
        "24.125,90,-1,10,20,30,40\n"
        "24.125,90,0,11,21,31,41\n",
        encoding="utf-8",
    )

    series_by_channel = load_hfss_summary_pattern(
        csv_path,
        ["Tx1", "Tx2", "Rx1", "Rx2"],
        value_kind=PATTERN_KIND_PHASE,
    )

    assert list(series_by_channel) == ["Tx1", "Tx2", "Rx1", "Rx2"]
    assert np.allclose(series_by_channel["Tx1"].values, np.array([-1.0, 0.0]))
    assert np.allclose(series_by_channel["Tx2"].values, np.array([-1.0, 0.0]))
    assert np.allclose(series_by_channel["Rx1"].values, np.array([-1.0, 0.0]))
    assert np.allclose(series_by_channel["Rx2"].values, np.array([-1.0, 0.0]))
    for series in series_by_channel.values():
        assert series.values_at(np.array([0.0]))[0] == pytest.approx(0.0)


def test_loads_1t2r_summary_with_rx_only_columns(tmp_path) -> None:
    csv_path = tmp_path / "rx-only-phase-summary.csv"
    csv_path.write_text(
        '"Theta [deg]","RX phase A","RX phase B"\n'
        "-1,10,20\n"
        "0,12,24\n",
        encoding="utf-8",
    )

    series_by_channel = load_hfss_summary_pattern(
        csv_path,
        ["Tx1", "Rx1", "Rx2"],
        value_kind=PATTERN_KIND_PHASE,
    )

    assert list(series_by_channel) == ["Rx1", "Rx2"]
    assert series_by_channel["Rx1"].value_column == "RX phase A"
    assert series_by_channel["Rx2"].value_column == "RX phase B"
    assert np.allclose(series_by_channel["Rx1"].values, np.array([-2.0, 0.0]))
    assert np.allclose(series_by_channel["Rx2"].values, np.array([-4.0, 0.0]))


def test_loads_1t2r_summary_with_tx_and_rx_columns(tmp_path) -> None:
    csv_path = tmp_path / "tx-rx-phase-summary.csv"
    csv_path.write_text(
        '"Theta [deg]","TX phase","RX phase A","RX phase B"\n'
        "-1,5,10,15\n"
        "0,6,12,18\n",
        encoding="utf-8",
    )

    series_by_channel = load_hfss_summary_pattern(
        csv_path,
        ["Tx1", "Rx1", "Rx2"],
        value_kind=PATTERN_KIND_PHASE,
    )

    assert list(series_by_channel) == ["Tx1", "Rx1", "Rx2"]
    assert series_by_channel["Tx1"].value_column == "TX phase"
    assert series_by_channel["Rx1"].value_column == "RX phase A"
    assert series_by_channel["Rx2"].value_column == "RX phase B"
    assert np.allclose(series_by_channel["Tx1"].values, np.array([-1.0, 0.0]))
    assert np.allclose(series_by_channel["Rx1"].values, np.array([-2.0, 0.0]))
    assert np.allclose(series_by_channel["Rx2"].values, np.array([-3.0, 0.0]))


def test_loads_summary_as_virtual_channels_in_tx_rx_order(tmp_path) -> None:
    csv_path = tmp_path / "virtual-phase-summary.csv"
    csv_path.write_text(
        '"Theta [deg]","CH1","CH2","CH3","CH4"\n'
        "-1,10,20,30,40\n"
        "0,11,22,33,44\n",
        encoding="utf-8",
    )
    channel_names = virtual_channel_names(["Tx1", "Tx2"], ["Rx1", "Rx2"])

    series_by_channel = load_hfss_summary_pattern(
        csv_path,
        channel_names,
        value_kind=PATTERN_KIND_PHASE,
    )

    assert list(series_by_channel) == ["Tx1Rx1", "Tx1Rx2", "Tx2Rx1", "Tx2Rx2"]
    assert series_by_channel["Tx1Rx1"].value_column == "CH1"
    assert series_by_channel["Tx1Rx2"].value_column == "CH2"
    assert series_by_channel["Tx2Rx1"].value_column == "CH3"
    assert series_by_channel["Tx2Rx2"].value_column == "CH4"
    assert np.allclose(series_by_channel["Tx2Rx2"].values, np.array([-4.0, 0.0]))


def test_loads_xlsx_summary_as_virtual_channels_in_tx_rx_order(tmp_path) -> None:
    path = tmp_path / "virtual-phase-summary.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Theta [deg]", "CH1", "CH2", "CH3", "CH4"])
    sheet.append([-1, 10, 20, 30, 40])
    sheet.append([0, 11, 22, 33, 44])
    workbook.save(path)
    channel_names = virtual_channel_names(["Tx1", "Tx2"], ["Rx1", "Rx2"])

    series_by_channel = load_hfss_summary_pattern(
        path,
        channel_names,
        value_kind=PATTERN_KIND_PHASE,
    )

    assert list(series_by_channel) == ["Tx1Rx1", "Tx1Rx2", "Tx2Rx1", "Tx2Rx2"]
    assert series_by_channel["Tx1Rx1"].value_column == "CH1"
    assert series_by_channel["Tx1Rx2"].value_column == "CH2"
    assert series_by_channel["Tx2Rx1"].value_column == "CH3"
    assert series_by_channel["Tx2Rx2"].value_column == "CH4"
    assert np.allclose(series_by_channel["Tx1Rx1"].values, np.array([-1.0, 0.0]))
    assert np.allclose(series_by_channel["Tx2Rx2"].values, np.array([-4.0, 0.0]))


def test_loads_xlsx_summary_without_openpyxl(tmp_path, monkeypatch) -> None:
    path = tmp_path / "virtual-phase-summary.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Theta [deg]", "CH1", "CH2"])
    sheet.append([-1, 10, 20])
    sheet.append([0, 11, 22])
    workbook.save(path)
    real_import = builtins.__import__

    def block_openpyxl(name, *args, **kwargs):  # noqa: ANN001
        if name == "openpyxl" or name.startswith("openpyxl."):
            raise ImportError("blocked for fallback test")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", block_openpyxl)

    series_by_channel = load_hfss_summary_pattern(
        path,
        ["Tx1Rx1", "Tx1Rx2"],
        value_kind=PATTERN_KIND_PHASE,
    )

    assert list(series_by_channel) == ["Tx1Rx1", "Tx1Rx2"]
    assert np.allclose(series_by_channel["Tx1Rx1"].values, np.array([-1.0, 0.0]))
    assert np.allclose(series_by_channel["Tx1Rx2"].values, np.array([-2.0, 0.0]))


def test_summary_explicit_virtual_headers_map_without_column_count_guess(tmp_path) -> None:
    csv_path = tmp_path / "virtual-named-phase-summary.csv"
    csv_path.write_text(
        '"Theta [deg]","phase_Tx1Rx2","Tx1/Rx1 phase"\n'
        "-1,10,20\n"
        "0,12,24\n",
        encoding="utf-8",
    )

    series_by_channel = load_hfss_summary_pattern(
        csv_path,
        ["Tx1", "Rx1", "Rx2"],
        value_kind=PATTERN_KIND_PHASE,
        virtual_channel_names=virtual_channel_names(["Tx1"], ["Rx1", "Rx2"]),
    )

    assert list(series_by_channel) == ["Tx1Rx2", "Tx1Rx1"]
    assert series_by_channel["Tx1Rx2"].value_column == "phase_Tx1Rx2"
    assert series_by_channel["Tx1Rx1"].value_column == "Tx1/Rx1 phase"


def test_loads_hfss_phase_summary_from_complex_values(tmp_path) -> None:
    csv_path = tmp_path / "complex-phase-summary.csv"
    csv_path.write_text(
        '"Freq [GHz]","Phi [deg]","Theta [deg]","E0","E1"\n'
        "24.125,90,-1,1+0j,0+1j\n"
        "24.125,90,0,0+1j,-1+0j\n",
        encoding="utf-8",
    )

    series_by_channel = load_hfss_summary_pattern(
        csv_path,
        ["Tx1", "Rx1"],
        value_kind=PATTERN_KIND_PHASE,
    )

    assert np.allclose(series_by_channel["Tx1"].values, np.array([-90.0, 0.0]))
    assert np.allclose(series_by_channel["Rx1"].values, np.array([-90.0, 0.0]))


def test_loads_hfss_phase_series_from_polar_complex_values(tmp_path) -> None:
    csv_path = tmp_path / "polar-phase.csv"
    csv_path.write_text(
        "Theta [deg],Phase\n"
        "-1,1∠10\n"
        "0,2∠40\n",
        encoding="utf-8",
    )

    series = load_hfss_pattern_series(
        csv_path,
        value_kind=PATTERN_KIND_PHASE,
    )

    assert np.allclose(series.values, np.array([-30.0, 0.0]))


def test_loads_hfss_phase_summary_from_xlsx_complex_values(tmp_path) -> None:
    path = tmp_path / "complex-phase-summary.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Freq [GHz]", "Phi [deg]", "Theta [deg]", "E0", "E1"])
    sheet.append([24.125, 90, -1, "1+0j", "0+1j"])
    sheet.append([24.125, 90, 0, "0+1j", "-1+0j"])
    workbook.save(path)

    series_by_channel = load_hfss_summary_pattern(
        path,
        ["Tx1", "Rx1"],
        value_kind=PATTERN_KIND_PHASE,
    )

    assert np.allclose(series_by_channel["Tx1"].values, np.array([-90.0, 0.0]))
    assert np.allclose(series_by_channel["Rx1"].values, np.array([-90.0, 0.0]))


def test_phase_series_unwraps_when_interpolating_across_wrap() -> None:
    phase_series = PatternSeries(
        name="phase",
        source_path="phase.csv",
        angle_column="Theta",
        value_column="phase",
        value_kind=PATTERN_KIND_PHASE,
        angles_deg=np.array([-1.0, 1.0]),
        values=np.array([179.0, -179.0]),
    )

    assert phase_series.values_at(np.array([0.0]))[0] == pytest.approx(180.0)


def test_pattern_series_clamps_angles_outside_csv_range() -> None:
    phase_series = PatternSeries(
        name="phase",
        source_path="phase.csv",
        angle_column="Theta",
        value_column="phase",
        value_kind=PATTERN_KIND_PHASE,
        angles_deg=np.array([-80.0, 0.0, 80.0]),
        values=np.array([-20.0, 0.0, 40.0]),
    )

    assert np.allclose(
        phase_series.values_at(np.array([-90.0, -40.0, 0.0, 40.0, 90.0])),
        np.array([-20.0, -10.0, 0.0, 20.0, 40.0]),
    )


def test_channel_pattern_weights_allow_partial_angle_coverage() -> None:
    phase_series = PatternSeries(
        name="phase",
        source_path="phase.xlsx",
        angle_column="Theta",
        value_column="RawData_Tx0_Rx0",
        value_kind=PATTERN_KIND_PHASE,
        angles_deg=np.array([-80.0, 0.0, 80.0]),
        values=np.array([-30.0, 0.0, 30.0]),
    )
    channel_patterns = ChannelPatternSet()
    channel_patterns.set_series(
        "Rx1",
        PATTERN_KIND_PHASE,
        PATTERN_PLANE_HORIZONTAL,
        phase_series,
    )

    weights = channel_patterns.complex_weights(
        ["Rx1"],
        np.array([-90.0, 0.0, 90.0]),
        np.array([0.0, 0.0, 0.0]),
    )

    assert weights.shape == (1, 3)
    assert weights[0, 0] == pytest.approx(np.exp(-1j * np.radians(30.0)))
    assert weights[0, 1] == pytest.approx(1.0 + 0.0j)
    assert weights[0, 2] == pytest.approx(np.exp(1j * np.radians(30.0)))


def test_element_pattern_axes_can_be_swapped(tmp_path) -> None:
    csv_path = tmp_path / "pattern.csv"
    csv_path.write_text(
        "Theta [deg],Phi0,Phi90\n"
        "-1,-10,-20\n"
        "0,0,-5\n"
        "1,-10,-20\n",
        encoding="utf-8",
    )

    pattern = load_element_pattern(csv_path)
    swapped = pattern.swapped_axes()

    assert swapped.horizontal_column == pattern.elevation_column
    assert swapped.elevation_column == pattern.horizontal_column
    assert np.array_equal(swapped.horizontal_gain_db, pattern.elevation_gain_db)
    assert np.array_equal(swapped.elevation_gain_db, pattern.horizontal_gain_db)


def test_single_gain_column_is_reused_as_horizontal_only(tmp_path) -> None:
    csv_path = tmp_path / "pattern.csv"
    csv_path.write_text("Angle(deg),Gain(dBi)\n-90,-30\n0,0\n90,-30\n", encoding="utf-8")

    pattern = load_element_pattern(csv_path)

    assert pattern.horizontal_column == "Gain(dBi)"
    assert pattern.elevation_column is None
    assert pattern.normalized_horizontal_gain_db_at(np.array([0.0]))[0] == 0.0
    assert pattern.normalized_elevation_gain_db_at(np.array([90.0]))[0] == -30.0


def test_element_pattern_weights_array_factor_grid() -> None:
    array = AntennaArray.from_xy(tx_x=[0], tx_y=[0], rx_x=[0], rx_y=[0])
    pattern = ElementPattern(
        name="test",
        source_path="test.csv",
        angle_column="Angle",
        horizontal_column="Horizontal",
        elevation_column=None,
        angles_deg=np.array([-75.0, 0.0, 75.0]),
        horizontal_gain_db=np.array([-20.0, 0.0, -20.0]),
    )

    af_db, azimuths, elevations, _metrics = calculate_metrics_and_psf(
        array,
        tx_pattern=pattern,
    )

    center_el = int(np.argmin(np.abs(elevations)))
    center_az = int(np.argmin(np.abs(azimuths)))
    edge_az = int(np.argmax(azimuths))

    assert af_db[center_el, center_az] == 0.0
    assert af_db[center_el, edge_az] == -20.0


def test_physical_channel_phase_pattern_weights_virtual_channels() -> None:
    array = AntennaArray.from_xy(tx_x=[0], tx_y=[0], rx_x=[0, 1], rx_y=[0, 0])
    phase_series = PatternSeries(
        name="phase",
        source_path="phase.csv",
        angle_column="Theta",
        value_column="phase",
        value_kind=PATTERN_KIND_PHASE,
        angles_deg=np.array([-90.0, 0.0, 90.0]),
        values=np.array([180.0, 180.0, 180.0]),
    )
    channel_patterns = ChannelPatternSet()
    channel_patterns.set_series(
        "Rx2",
        PATTERN_KIND_PHASE,
        PATTERN_PLANE_HORIZONTAL,
        phase_series,
    )

    af_db, azimuths, elevations, _metrics = calculate_metrics_and_psf(
        array,
        channel_patterns=channel_patterns,
    )

    center_el = int(np.argmin(np.abs(elevations)))
    center_az = int(np.argmin(np.abs(azimuths)))

    assert af_db[center_el, center_az] < -100.0


def test_virtual_channel_phase_pattern_weights_matching_virtual_channel() -> None:
    array = AntennaArray.from_xy(tx_x=[0], tx_y=[0], rx_x=[0, 1], rx_y=[0, 0])
    phase_series = PatternSeries(
        name="phase",
        source_path="phase.csv",
        angle_column="Theta",
        value_column="phase",
        value_kind=PATTERN_KIND_PHASE,
        angles_deg=np.array([-90.0, 0.0, 90.0]),
        values=np.array([180.0, 180.0, 180.0]),
    )
    channel_patterns = ChannelPatternSet()
    channel_patterns.set_series(
        "Tx1Rx2",
        PATTERN_KIND_PHASE,
        PATTERN_PLANE_HORIZONTAL,
        phase_series,
    )

    af_db, azimuths, elevations, _metrics = calculate_metrics_and_psf(
        array,
        channel_patterns=channel_patterns,
    )

    center_el = int(np.argmin(np.abs(elevations)))
    center_az = int(np.argmin(np.abs(azimuths)))
    assert af_db[center_el, center_az] < -100.0


def test_dbf_uses_channel_patterns_on_true_signal_not_scan_dictionary() -> None:
    array = AntennaArray.from_xy(tx_x=[0], tx_y=[0], rx_x=[0, 1], rx_y=[0, 0])
    phase_series = PatternSeries(
        name="phase",
        source_path="phase.csv",
        angle_column="Theta",
        value_column="phase",
        value_kind=PATTERN_KIND_PHASE,
        angles_deg=np.array([-90.0, 0.0, 90.0]),
        values=np.array([180.0, 180.0, 180.0]),
    )
    channel_patterns = ChannelPatternSet()
    channel_patterns.set_series(
        "Rx2",
        PATTERN_KIND_PHASE,
        PATTERN_PLANE_HORIZONTAL,
        phase_series,
    )

    angles, spectrum_db = dbf_azimuth_spectrum(
        array,
        true_angle_deg=0.0,
        angles_deg=np.array([-90.0, 0.0, 90.0]),
        channel_patterns=channel_patterns,
    )

    center_index = int(np.where(angles == 0.0)[0][0])
    assert spectrum_db[center_index] < -100.0
    assert np.max(spectrum_db) == pytest.approx(0.0)


def test_dbf_uses_hfss_phase_as_signal_without_extra_ideal_phase() -> None:
    array = AntennaArray.from_xy(
        tx_x=[0],
        tx_y=[0],
        rx_x=[0, 1, 2],
        rx_y=[0, 0, 0],
    )
    angles_deg = np.array([-30.0, 0.0, 30.0])
    channel_patterns = ChannelPatternSet()
    for rx_name, position in (("Rx2", 1.0), ("Rx3", 2.0)):
        phase_series = PatternSeries(
            name=f"{rx_name}-phase",
            source_path=f"{rx_name}.csv",
            angle_column="Theta",
            value_column="phase",
            value_kind=PATTERN_KIND_PHASE,
            angles_deg=angles_deg,
            values=180.0 * position * np.sin(np.radians(angles_deg)),
        )
        channel_patterns.set_series(
            rx_name,
            PATTERN_KIND_PHASE,
            PATTERN_PLANE_HORIZONTAL,
            phase_series,
        )

    scan_angles, spectrum_db = dbf_azimuth_spectrum(
        array,
        true_angle_deg=30.0,
        angles_deg=angles_deg,
        channel_patterns=channel_patterns,
    )

    peak_index = int(np.argmax(spectrum_db))
    center_index = int(np.where(scan_angles == 0.0)[0][0])
    assert scan_angles[peak_index] == pytest.approx(30.0)
    assert spectrum_db[peak_index] == pytest.approx(0.0)
    assert spectrum_db[center_index] < -3.0


def test_dbf_auto_selects_steering_sign_for_hfss_phase_convention() -> None:
    array = AntennaArray.from_xy(
        tx_x=[0],
        tx_y=[0],
        rx_x=[0, 1, 2],
        rx_y=[0, 0, 0],
    )
    angles_deg = np.array([-30.0, 0.0, 30.0])
    channel_patterns = ChannelPatternSet()
    for rx_name, position in (("Rx2", 1.0), ("Rx3", 2.0)):
        phase_series = PatternSeries(
            name=f"{rx_name}-phase",
            source_path=f"{rx_name}.csv",
            angle_column="Theta",
            value_column="phase",
            value_kind=PATTERN_KIND_PHASE,
            angles_deg=angles_deg,
            values=-180.0 * position * np.sin(np.radians(angles_deg)),
        )
        channel_patterns.set_series(
            rx_name,
            PATTERN_KIND_PHASE,
            PATTERN_PLANE_HORIZONTAL,
            phase_series,
        )

    scan_angles, spectrum_db = dbf_azimuth_spectrum(
        array,
        true_angle_deg=30.0,
        angles_deg=angles_deg,
        channel_patterns=channel_patterns,
    )

    peak_index = int(np.argmax(spectrum_db))
    assert scan_angles[peak_index] == pytest.approx(30.0)


def test_pattern_cut_metrics_reports_peak_and_beamwidths() -> None:
    angles = np.array([-4.0, -2.0, 0.0, 2.0, 4.0])
    gain = np.array([-8.0, -3.0, 1.0, -3.0, -8.0])

    metrics = pattern_cut_metrics(angles, gain)

    assert metrics.peak_angle_deg == 0.0
    assert metrics.peak_gain_dbi == 1.0
    assert metrics.beamwidth_3db_deg == 3.0
    assert metrics.beamwidth_6db_deg == 5.6
    assert format_pattern_cut_metrics(metrics) == (
        "Peak 1.00 dBi @ 0.0° | 3dB BW 3.0° | 6dB BW 5.6°"
    )


def test_pattern_cut_metrics_returns_none_for_open_beamwidth() -> None:
    angles = np.array([-1.0, 0.0, 1.0])
    gain = np.array([-1.0, 0.0, -1.0])

    metrics = pattern_cut_metrics(angles, gain)

    assert metrics.beamwidth_3db_deg is None
    assert metrics.beamwidth_6db_deg is None
    assert "3dB BW N/A" in format_pattern_cut_metrics(metrics)
