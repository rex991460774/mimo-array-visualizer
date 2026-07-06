from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path

import numpy as np

from .element_pattern import ChannelPatternSet, virtual_channel_names
from .geometry import AntennaArray
from .table_io import read_xlsx_rows


DBF_DICT_IDEAL = "ideal"
DBF_DICT_IDEAL_REVERSED = "ideal_reversed"
DBF_DICT_CHANNEL_PATTERN = "channel_pattern"
# Legacy state value. The UI now folds this behavior into DBF_DICT_CHANNEL_PATTERN.
DBF_DICT_CHANNEL_PATTERN_ZERO_REF = "channel_pattern_zero_ref"
DBF_DICT_CUSTOM = "custom"

DBF_DICTIONARY_MODES = (
    DBF_DICT_IDEAL,
    DBF_DICT_IDEAL_REVERSED,
    DBF_DICT_CHANNEL_PATTERN,
    DBF_DICT_CUSTOM,
)

CHANNEL_MODE_VIRTUAL = "virtual"
CHANNEL_MODE_PHYSICAL = "physical"
DBF_DICTIONARY_QUALITY_OK = "ok"
DBF_DICTIONARY_QUALITY_WARNING = "warning"
DBF_DICTIONARY_QUALITY_DANGER = "danger"
DBF_DICTIONARY_COMPETITOR_MARGIN_DB = 0.5
DBF_DICTIONARY_FAR_ANGLE_SEPARATION_DEG = 10.0
DBF_DICTIONARY_FAR_SPATIAL_SEPARATION = float(
    np.sin(np.radians(DBF_DICTIONARY_FAR_ANGLE_SEPARATION_DEG))
)
DBF_DICTIONARY_ROW_NORM_EPS = 1e-12


@dataclass(frozen=True)
class DbfDictionaryQualityReport:
    severity: str
    row_count: int
    channel_count: int
    zero_norm_rows: int
    competitor_ambiguous_rows: int
    non_neighbor_effective_rank: int
    non_neighbor_rank_rows: int


@dataclass(frozen=True)
class DbfDictionaryTable:
    source_path: str
    angle_column: str
    azimuth_column: str | None
    elevation_column: str | None
    channel_mode: str
    value_format: str
    column_names: tuple[str, ...]
    angles_deg: np.ndarray
    azimuths_deg: np.ndarray
    elevations_deg: np.ndarray
    values: np.ndarray

    @property
    def display_name(self) -> str:
        return Path(self.source_path).name

    @property
    def is_2d(self) -> bool:
        return self.azimuth_column is not None and self.elevation_column is not None

    def response_matrix(
        self,
        array: AntennaArray,
        azimuth_deg: np.ndarray,
        elevation_deg: np.ndarray,
        axis: str,
        *,
        zero_phase_calibrated: bool = False,
        phase_reversed: bool = False,
    ) -> np.ndarray:
        azimuths, elevations = np.broadcast_arrays(
            np.asarray(azimuth_deg, dtype=float),
            np.asarray(elevation_deg, dtype=float),
        )
        flat_az = azimuths.ravel()
        flat_el = elevations.ravel()
        rows = (
            self._nearest_rows(flat_az, flat_el)
            if self.is_2d
            else self._interpolated_rows(flat_az if axis == "azimuth" else flat_el)
        )
        rows = self._adjust_rows(
            rows,
            zero_phase_calibrated=zero_phase_calibrated,
            phase_reversed=phase_reversed,
        )
        virtual = self._virtual_rows(array, rows)
        return virtual.reshape(*azimuths.shape, len(array.tx) * len(array.rx))

    def _adjust_rows(
        self,
        rows: np.ndarray,
        *,
        zero_phase_calibrated: bool,
        phase_reversed: bool,
    ) -> np.ndarray:
        adjusted = np.asarray(rows, dtype=complex)
        if zero_phase_calibrated:
            adjusted = _remove_reference_phase(adjusted, self._zero_reference_row())
        if phase_reversed:
            adjusted = np.conjugate(adjusted)
        return adjusted

    def _zero_reference_row(self) -> np.ndarray:
        if self.is_2d:
            return self._nearest_rows(np.asarray([0.0]), np.asarray([0.0]))[0]
        return self._interpolated_rows(np.asarray([0.0]))[0]

    def _interpolated_rows(self, angles_deg: np.ndarray) -> np.ndarray:
        query_angles = np.asarray(angles_deg, dtype=float)
        amplitudes = np.abs(self.values)
        phases = np.unwrap(np.angle(self.values), axis=0)
        interpolated_amp = np.empty((len(query_angles), self.values.shape[1]), dtype=float)
        interpolated_phase = np.empty_like(interpolated_amp)
        for column in range(self.values.shape[1]):
            interpolated_amp[:, column] = np.interp(
                query_angles,
                self.angles_deg,
                amplitudes[:, column],
                left=amplitudes[0, column],
                right=amplitudes[-1, column],
            )
            interpolated_phase[:, column] = np.interp(
                query_angles,
                self.angles_deg,
                phases[:, column],
                left=phases[0, column],
                right=phases[-1, column],
            )
        return interpolated_amp * np.exp(1j * interpolated_phase)

    def _nearest_rows(self, azimuth_deg: np.ndarray, elevation_deg: np.ndarray) -> np.ndarray:
        points = np.column_stack([self.azimuths_deg, self.elevations_deg])
        queries = np.column_stack([azimuth_deg, elevation_deg])
        indices = []
        for query in queries:
            distance = np.sum((points - query[None, :]) ** 2, axis=1)
            indices.append(int(np.argmin(distance)))
        return self.values[np.asarray(indices, dtype=int)]

    def _virtual_rows(self, array: AntennaArray, rows: np.ndarray) -> np.ndarray:
        virtual_count = len(array.tx) * len(array.rx)
        if self.channel_mode == CHANNEL_MODE_VIRTUAL:
            if rows.shape[1] != virtual_count:
                raise ValueError(
                    f"DBF dictionary has {rows.shape[1]} virtual channel columns, "
                    f"but the current layout needs {virtual_count}."
                )
            return _reorder_virtual_columns(array, rows, self.column_names)

        tx_count = len(array.tx)
        rx_count = len(array.rx)
        physical_counts = {tx_count + rx_count}
        if tx_count == 1:
            physical_counts.add(rx_count)
        if rows.shape[1] not in physical_counts:
            expected_text = " or ".join(str(count) for count in sorted(physical_counts))
            raise ValueError(
                f"DBF dictionary has {rows.shape[1]} physical channel columns, "
                f"but the current layout needs {expected_text}."
            )
        tx_rows, rx_rows = _split_physical_columns(array, rows, self.column_names)
        virtual = tx_rows[:, :, None] * rx_rows[:, None, :]
        return virtual.reshape(rows.shape[0], virtual_count)


@dataclass(frozen=True)
class DbfDictionaryConfig:
    mode: str = DBF_DICT_IDEAL
    custom_azimuth_table: DbfDictionaryTable | None = None
    custom_elevation_table: DbfDictionaryTable | None = None
    custom_table: DbfDictionaryTable | None = None
    custom_phase_reversed: bool = False
    custom_zero_phase_calibrated: bool = False

    def __post_init__(self) -> None:
        if self.mode == DBF_DICT_CHANNEL_PATTERN_ZERO_REF:
            object.__setattr__(self, "mode", DBF_DICT_CHANNEL_PATTERN)
        if self.custom_table is None:
            return
        if self.custom_azimuth_table is None:
            object.__setattr__(self, "custom_azimuth_table", self.custom_table)
        if self.custom_elevation_table is None:
            object.__setattr__(self, "custom_elevation_table", self.custom_table)

    @property
    def uses_auto_ideal_sign(self) -> bool:
        return self.mode == DBF_DICT_IDEAL

    @property
    def display_name(self) -> str:
        if self.mode == DBF_DICT_CUSTOM:
            names = []
            if self.custom_azimuth_table is not None:
                names.append(f"az={self.custom_azimuth_table.display_name}")
            if self.custom_elevation_table is not None:
                names.append(f"el={self.custom_elevation_table.display_name}")
            if names:
                return f"custom: {', '.join(names)}"
        return self.mode

    def custom_table_for_axis(self, axis: str) -> DbfDictionaryTable | None:
        if axis == "elevation":
            return self.custom_elevation_table
        return self.custom_azimuth_table

    def scan_matrix(
        self,
        array: AntennaArray,
        angles_deg: np.ndarray,
        axis: str,
        channel_patterns: ChannelPatternSet | None = None,
    ) -> np.ndarray:
        angles = np.asarray(angles_deg, dtype=float)
        if self.mode == DBF_DICT_CUSTOM:
            response = self._custom_axis_response(array, angles, axis)
            return np.conjugate(response)
        zeros = np.zeros_like(angles)
        azimuths = angles if axis == "azimuth" else zeros
        elevations = angles if axis == "elevation" else zeros
        return self.scan_matrix_2d(
            array,
            azimuths,
            elevations,
            axis=axis,
            channel_patterns=channel_patterns,
        )

    def scan_matrix_2d(
        self,
        array: AntennaArray,
        azimuth_deg: np.ndarray,
        elevation_deg: np.ndarray,
        axis: str = "azimuth",
        channel_patterns: ChannelPatternSet | None = None,
    ) -> np.ndarray:
        azimuths, elevations = np.broadcast_arrays(
            np.asarray(azimuth_deg, dtype=float),
            np.asarray(elevation_deg, dtype=float),
        )
        flat_az = azimuths.ravel()
        flat_el = elevations.ravel()

        if self.mode == DBF_DICT_IDEAL_REVERSED:
            return _ideal_scan_matrix(array, flat_az, flat_el, steering_sign=1)
        if self.mode == DBF_DICT_CHANNEL_PATTERN:
            response = _channel_pattern_response(
                array, flat_az, flat_el, channel_patterns=channel_patterns
            )
            reference = _channel_pattern_response(
                array,
                np.zeros_like(flat_az),
                np.zeros_like(flat_el),
                channel_patterns=channel_patterns,
            )
            response = _remove_reference_phase(response, reference)
            return np.conjugate(response)
        if self.mode == DBF_DICT_CUSTOM:
            response = self._custom_2d_response(array, flat_az, flat_el, axis=axis)
            return np.conjugate(response)
        return _ideal_scan_matrix(array, flat_az, flat_el, steering_sign=-1)

    def _custom_axis_response(
        self,
        array: AntennaArray,
        angles_deg: np.ndarray,
        axis: str,
    ) -> np.ndarray:
        table = self.custom_table_for_axis(axis)
        if table is None:
            zeros = np.zeros_like(angles_deg, dtype=float)
            azimuths = angles_deg if axis == "azimuth" else zeros
            elevations = angles_deg if axis == "elevation" else zeros
            return _ideal_scan_matrix(array, azimuths, elevations, steering_sign=1)
        zeros = np.zeros_like(angles_deg, dtype=float)
        azimuths = angles_deg if axis == "azimuth" else zeros
        elevations = angles_deg if axis == "elevation" else zeros
        return table.response_matrix(
            array,
            azimuths,
            elevations,
            axis=axis,
            zero_phase_calibrated=self.custom_zero_phase_calibrated,
            phase_reversed=self.custom_phase_reversed,
        )

    def _custom_2d_response(
        self,
        array: AntennaArray,
        azimuth_deg: np.ndarray,
        elevation_deg: np.ndarray,
        axis: str,
    ) -> np.ndarray:
        az_table = self.custom_azimuth_table
        el_table = self.custom_elevation_table
        if az_table is None or el_table is None:
            return self._custom_separable_2d_response(
                array,
                azimuth_deg,
                elevation_deg,
            )
        if az_table is el_table and az_table.is_2d:
            return az_table.response_matrix(
                array,
                azimuth_deg,
                elevation_deg,
                axis=axis,
                zero_phase_calibrated=self.custom_zero_phase_calibrated,
                phase_reversed=self.custom_phase_reversed,
            )

        return self._custom_separable_2d_response(array, azimuth_deg, elevation_deg)

    def _custom_separable_2d_response(
        self,
        array: AntennaArray,
        azimuth_deg: np.ndarray,
        elevation_deg: np.ndarray,
    ) -> np.ndarray:
        az_table = self.custom_azimuth_table
        el_table = self.custom_elevation_table
        zeros = np.zeros_like(azimuth_deg, dtype=float)
        az_response = _axis_response_or_ideal(
            array,
            az_table,
            azimuth_deg,
            zeros,
            axis="azimuth",
            zero_phase_calibrated=self.custom_zero_phase_calibrated,
            phase_reversed=self.custom_phase_reversed,
        )
        el_response = _axis_response_or_ideal(
            array,
            el_table,
            zeros,
            elevation_deg,
            axis="elevation",
            zero_phase_calibrated=self.custom_zero_phase_calibrated,
            phase_reversed=self.custom_phase_reversed,
        )
        az_reference = _axis_response_or_ideal(
            array,
            az_table,
            zeros,
            zeros,
            axis="azimuth",
            zero_phase_calibrated=self.custom_zero_phase_calibrated,
            phase_reversed=self.custom_phase_reversed,
        )
        el_reference = _axis_response_or_ideal(
            array,
            el_table,
            zeros,
            zeros,
            axis="elevation",
            zero_phase_calibrated=self.custom_zero_phase_calibrated,
            phase_reversed=self.custom_phase_reversed,
        )
        az_delta = _safe_complex_divide(az_response, az_reference)
        el_delta = _safe_complex_divide(el_response, el_reference)
        if az_table is None and el_table is None:
            base_response = np.ones_like(az_reference, dtype=complex)
        elif az_table is None:
            base_response = el_reference
        elif el_table is None:
            base_response = az_reference
        else:
            base_response = _mean_complex_reference(az_reference, el_reference)
        return base_response * az_delta * el_delta


def load_dbf_dictionary_table(
    path: str | Path,
    array: AntennaArray,
    channel_mode: str | None = None,
) -> DbfDictionaryTable:
    source_path = Path(path)
    header, raw_rows = _read_table(source_path)
    if len(header) < 2 or not raw_rows:
        raise ValueError("DBF dictionary file must contain a header and data rows.")

    angle_index, azimuth_index, elevation_index = _angle_column_indices(header)
    angle_indices = {
        index
        for index in (angle_index, azimuth_index, elevation_index)
        if index is not None
    }
    data_indices = _dbf_data_column_indices(header, angle_indices, angle_index)
    if not data_indices:
        raise ValueError("DBF dictionary file must contain channel columns.")

    angles = []
    azimuths = []
    elevations = []
    value_rows = []
    raw_value_strings: list[list[str]] = []
    for row_number, row in enumerate(raw_rows, start=2):
        if max([*angle_indices, *data_indices]) >= len(row):
            continue
        try:
            angle = _parse_float(row[angle_index]) if angle_index is not None else 0.0
            azimuth = (
                _parse_float(row[azimuth_index])
                if azimuth_index is not None
                else angle
            )
            elevation = (
                _parse_float(row[elevation_index])
                if elevation_index is not None
                else 0.0
            )
            values = [_parse_complex_or_float(row[index]) for index in data_indices]
        except ValueError as exc:
            raise ValueError(f"Invalid DBF dictionary value on row {row_number}.") from exc
        if np.isfinite(angle) and np.isfinite(azimuth) and np.isfinite(elevation):
            angles.append(angle)
            azimuths.append(azimuth)
            elevations.append(elevation)
            value_rows.append(values)
            raw_value_strings.append([row[index] for index in data_indices])

    if len(value_rows) < 1:
        raise ValueError("DBF dictionary file does not contain valid numeric rows.")

    values = np.asarray(value_rows, dtype=complex)
    raw_has_complex = any(_looks_complex(cell) for row in raw_value_strings for cell in row)
    value_format = "complex" if raw_has_complex else "phase_deg"
    if value_format == "phase_deg":
        values = np.exp(1j * np.radians(values.real))
    channel_names = tuple(header[index].strip() for index in data_indices)
    resolved_channel_mode = _infer_channel_mode(
        array,
        channel_names,
        values.shape[1],
        preferred_mode=channel_mode,
    )

    order = np.lexsort((np.asarray(elevations, dtype=float), np.asarray(azimuths, dtype=float)))
    if azimuth_index is None and elevation_index is None:
        order = np.argsort(np.asarray(angles, dtype=float))
    sorted_angles = np.asarray(angles, dtype=float)[order]
    sorted_azimuths = np.asarray(azimuths, dtype=float)[order]
    sorted_elevations = np.asarray(elevations, dtype=float)[order]
    sorted_values = values[order]
    return DbfDictionaryTable(
        source_path=str(source_path),
        angle_column=header[angle_index] if angle_index is not None else "",
        azimuth_column=header[azimuth_index] if azimuth_index is not None else None,
        elevation_column=header[elevation_index] if elevation_index is not None else None,
        channel_mode=resolved_channel_mode,
        value_format=value_format,
        column_names=channel_names,
        angles_deg=sorted_angles,
        azimuths_deg=sorted_azimuths,
        elevations_deg=sorted_elevations,
        values=sorted_values,
    )


def dictionary_phase_preview(matrix: np.ndarray) -> np.ndarray:
    return np.angle(np.asarray(matrix, dtype=complex), deg=True)


def dictionary_quality_report(
    matrix: np.ndarray,
    angles_deg: np.ndarray | None = None,
) -> DbfDictionaryQualityReport:
    rows = np.asarray(matrix, dtype=complex)
    if rows.ndim != 2:
        rows = np.atleast_2d(rows)
    row_count, channel_count = rows.shape if rows.ndim == 2 else (0, 0)
    if row_count == 0 or channel_count == 0:
        return DbfDictionaryQualityReport(
            severity=DBF_DICTIONARY_QUALITY_DANGER,
            row_count=row_count,
            channel_count=channel_count,
            zero_norm_rows=row_count,
            competitor_ambiguous_rows=0,
            non_neighbor_effective_rank=0,
            non_neighbor_rank_rows=0,
        )

    row_norms = np.linalg.norm(rows, axis=1)
    valid_mask = row_norms > DBF_DICTIONARY_ROW_NORM_EPS
    valid_rows = rows[valid_mask]
    zero_norm_rows = int(row_count - int(np.count_nonzero(valid_mask)))
    if valid_rows.shape[0] < 2:
        return DbfDictionaryQualityReport(
            severity=DBF_DICTIONARY_QUALITY_DANGER,
            row_count=row_count,
            channel_count=channel_count,
            zero_norm_rows=zero_norm_rows,
            competitor_ambiguous_rows=0,
            non_neighbor_effective_rank=int(valid_rows.shape[0]),
            non_neighbor_rank_rows=int(valid_rows.shape[0]),
        )

    normalized = valid_rows / row_norms[valid_mask, None]
    correlation = np.abs(normalized @ np.conjugate(normalized.T))
    spectra_db = 20.0 * np.log10(
        np.clip(correlation, DBF_DICTIONARY_ROW_NORM_EPS, 1.0)
    )
    competitor_ambiguous_rows = sum(
        1
        for row_index, spectrum_db in enumerate(spectra_db)
        if _dictionary_peak_margin_db(spectrum_db, row_index)
        <= DBF_DICTIONARY_COMPETITOR_MARGIN_DB
    )

    representative_positions = _dictionary_representative_positions(
        valid_mask,
        row_count,
        angles_deg,
    )
    representative_rows = normalized[representative_positions]
    non_neighbor_effective_rank = _dictionary_effective_rank(representative_rows)
    non_neighbor_rank_rows = int(representative_rows.shape[0])

    severity = DBF_DICTIONARY_QUALITY_OK
    if (
        zero_norm_rows > 0
        or competitor_ambiguous_rows > 0
        or (non_neighbor_rank_rows > 1 and non_neighbor_effective_rank <= 1)
    ):
        severity = DBF_DICTIONARY_QUALITY_DANGER

    return DbfDictionaryQualityReport(
        severity=severity,
        row_count=row_count,
        channel_count=channel_count,
        zero_norm_rows=zero_norm_rows,
        competitor_ambiguous_rows=competitor_ambiguous_rows,
        non_neighbor_effective_rank=non_neighbor_effective_rank,
        non_neighbor_rank_rows=non_neighbor_rank_rows,
    )


def _dictionary_far_pair_mask(
    valid_mask: np.ndarray,
    row_count: int,
    angles_deg: np.ndarray | None,
) -> np.ndarray:
    valid_indices = np.flatnonzero(valid_mask)
    valid_coordinates = _dictionary_valid_spatial_coordinates(
        valid_mask,
        row_count,
        angles_deg,
    )
    if valid_coordinates is not None:
        return (
            np.abs(valid_coordinates[:, None] - valid_coordinates[None, :])
            >= DBF_DICTIONARY_FAR_SPATIAL_SEPARATION
        )
    index_delta = np.abs(valid_indices[:, None] - valid_indices[None, :])
    return index_delta > 1


def _dictionary_valid_spatial_coordinates(
    valid_mask: np.ndarray,
    row_count: int,
    angles_deg: np.ndarray | None,
) -> np.ndarray | None:
    if angles_deg is None:
        return None
    angles = np.asarray(angles_deg, dtype=float)
    if angles.shape[0] != row_count:
        return None
    return _dictionary_spatial_coordinates(angles[np.flatnonzero(valid_mask)])


def _dictionary_peak_margin_db(spectrum_db: np.ndarray, main_index: int) -> float:
    values = np.asarray(spectrum_db, dtype=float)
    if values.size == 0:
        return float("-inf")
    main_peak = float(values[main_index])
    left_bound, right_bound = _dictionary_main_lobe_bounds(values, main_index)
    competitor_indices = [
        index
        for index in _dictionary_local_peak_indices(values)
        if index < left_bound or index > right_bound
    ]
    if not competitor_indices:
        return float("inf")
    competitor_peak = max(float(values[index]) for index in competitor_indices)
    return main_peak - competitor_peak


def _dictionary_main_lobe_bounds(values_db: np.ndarray, main_index: int) -> tuple[int, int]:
    left = main_index
    while left > 0:
        current = float(values_db[left])
        prev_value = float(values_db[left - 1])
        next_value = float(values_db[left + 1]) if left + 1 < len(values_db) else current
        if (
            left < main_index
            and current <= prev_value
            and current <= next_value
            and (current < prev_value or current < next_value)
        ):
            break
        left -= 1

    right = main_index
    while right < len(values_db) - 1:
        current = float(values_db[right])
        prev_value = float(values_db[right - 1]) if right > 0 else current
        next_value = float(values_db[right + 1])
        if (
            right > main_index
            and current <= prev_value
            and current <= next_value
            and (current < prev_value or current < next_value)
        ):
            break
        right += 1
    return left, right


def _dictionary_local_peak_indices(values_db: np.ndarray) -> list[int]:
    if len(values_db) == 0:
        return []
    peaks: list[int] = []
    index = 0
    while index < len(values_db):
        start = index
        while (
            index + 1 < len(values_db)
            and abs(float(values_db[index + 1]) - float(values_db[start])) <= 1e-9
        ):
            index += 1
        end = index
        value = float(values_db[start])
        left = float(values_db[start - 1]) if start > 0 else float("-inf")
        right = float(values_db[end + 1]) if end + 1 < len(values_db) else float("-inf")
        if value >= left and value >= right and (value > left or value > right):
            peaks.append((start + end) // 2)
        index += 1
    return peaks


def _dictionary_representative_positions(
    valid_mask: np.ndarray,
    row_count: int,
    angles_deg: np.ndarray | None,
) -> np.ndarray:
    valid_indices = np.flatnonzero(valid_mask)
    if valid_indices.size == 0:
        return np.empty(0, dtype=int)
    if angles_deg is not None:
        angles = np.asarray(angles_deg, dtype=float)
        if angles.shape[0] == row_count:
            valid_coordinates = _dictionary_spatial_coordinates(angles[valid_indices])
            order = np.argsort(valid_coordinates)
            selected: list[int] = []
            last_coordinate: float | None = None
            for position in order:
                coordinate = float(valid_coordinates[position])
                if (
                    last_coordinate is None
                    or coordinate - last_coordinate
                    >= DBF_DICTIONARY_FAR_SPATIAL_SEPARATION
                ):
                    selected.append(int(position))
                    last_coordinate = coordinate
            return np.asarray(selected, dtype=int)

    selected = []
    last_index: int | None = None
    for position, original_index in enumerate(valid_indices):
        if last_index is None or int(original_index) - last_index > 1:
            selected.append(position)
            last_index = int(original_index)
    return np.asarray(selected, dtype=int)


def _dictionary_spatial_coordinates(angles_deg: np.ndarray) -> np.ndarray:
    return np.sin(np.radians(np.asarray(angles_deg, dtype=float)))


def _dictionary_effective_rank(rows: np.ndarray) -> int:
    matrix = np.asarray(rows, dtype=complex)
    if matrix.size == 0:
        return 0
    singular_values = np.linalg.svd(matrix, compute_uv=False)
    if singular_values.size == 0:
        return 0
    rank_tolerance = max(matrix.shape) * np.finfo(float).eps * float(singular_values[0])
    return int(np.count_nonzero(singular_values > rank_tolerance))


def _remove_reference_phase(
    rows: np.ndarray,
    reference: np.ndarray,
) -> np.ndarray:
    row_matrix = np.asarray(rows, dtype=complex)
    reference_phase = np.angle(np.asarray(reference, dtype=complex))
    phase_factor = np.exp(-1j * reference_phase)
    if phase_factor.ndim == 1:
        phase_factor = phase_factor[None, :]
    return row_matrix * phase_factor


def _axis_response_or_ideal(
    array: AntennaArray,
    table: DbfDictionaryTable | None,
    azimuth_deg: np.ndarray,
    elevation_deg: np.ndarray,
    axis: str,
    *,
    zero_phase_calibrated: bool = False,
    phase_reversed: bool = False,
) -> np.ndarray:
    if table is not None:
        return table.response_matrix(
            array,
            azimuth_deg,
            elevation_deg,
            axis=axis,
            zero_phase_calibrated=zero_phase_calibrated,
            phase_reversed=phase_reversed,
        )
    return _ideal_scan_matrix(array, azimuth_deg, elevation_deg, steering_sign=1)


def _safe_complex_divide(numerator: np.ndarray, denominator: np.ndarray) -> np.ndarray:
    denominator_abs = np.abs(denominator)
    return np.divide(
        numerator,
        denominator,
        out=np.ones_like(numerator, dtype=complex),
        where=denominator_abs > 0.0,
    )


def _mean_complex_reference(
    azimuth_reference: np.ndarray,
    elevation_reference: np.ndarray,
) -> np.ndarray:
    amplitude = np.sqrt(np.abs(azimuth_reference) * np.abs(elevation_reference))
    phase = 0.5 * (np.angle(azimuth_reference) + np.angle(elevation_reference))
    return amplitude * np.exp(1j * phase)


def _ideal_scan_matrix(
    array: AntennaArray,
    azimuth_deg: np.ndarray,
    elevation_deg: np.ndarray,
    steering_sign: int,
) -> np.ndarray:
    virtual_xy = array.virtual_xy()
    azimuth_rad = np.radians(azimuth_deg)
    elevation_rad = np.radians(elevation_deg)
    u = np.sin(azimuth_rad) * np.cos(elevation_rad)
    v = np.sin(elevation_rad)
    phase = steering_sign * np.pi * (
        u[:, None] * virtual_xy[None, :, 0]
        + v[:, None] * virtual_xy[None, :, 1]
    )
    return np.exp(1j * phase)


def _channel_pattern_response(
    array: AntennaArray,
    azimuth_deg: np.ndarray,
    elevation_deg: np.ndarray,
    channel_patterns: ChannelPatternSet | None,
) -> np.ndarray:
    if channel_patterns is None or channel_patterns.is_empty():
        return np.conjugate(_ideal_scan_matrix(array, azimuth_deg, elevation_deg, -1))
    tx_names = [point.name for point in array.tx]
    rx_names = [point.name for point in array.rx]
    tx_weights = channel_patterns.complex_weights(tx_names, azimuth_deg, elevation_deg)
    rx_weights = channel_patterns.complex_weights(rx_names, azimuth_deg, elevation_deg)
    virtual = tx_weights[:, None, :] * rx_weights[None, :, :]
    combined = np.moveaxis(virtual, -1, 0).reshape(
        len(azimuth_deg), len(tx_names) * len(rx_names)
    )
    direct_virtual = channel_patterns.complex_weights(
        virtual_channel_names(tx_names, rx_names),
        azimuth_deg,
        elevation_deg,
    ).T
    return combined * direct_virtual


def _read_table(path: Path) -> tuple[list[str], list[list[str]]]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            from openpyxl import load_workbook
        except ImportError:
            rows = read_xlsx_rows(path)
        else:
            workbook = load_workbook(path, data_only=True, read_only=True)
            sheet = workbook.active
            rows = [
                ["" if cell is None else str(cell).strip() for cell in row]
                for row in sheet.iter_rows(values_only=True)
                if row and any(cell is not None and str(cell).strip() for cell in row)
            ]
    else:
        delimiter = _detect_delimiter(path)
        with path.open("r", encoding="utf-8-sig", newline="") as file:
            rows = [
                [cell.strip() for cell in row]
                for row in csv.reader(file, delimiter=delimiter)
                if row and any(cell.strip() for cell in row)
            ]
    if len(rows) < 2:
        raise ValueError("DBF dictionary file must contain at least two rows.")
    return rows[0], rows[1:]


def _detect_delimiter(path: Path) -> str:
    if path.suffix.lower() == ".tsv":
        return "\t"
    sample = path.read_text(encoding="utf-8-sig", errors="ignore")[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        return ","
    return dialect.delimiter


def _angle_column_indices(header: list[str]) -> tuple[int | None, int | None, int | None]:
    normalized = [_normalize_header(value) for value in header]
    azimuth_index = _first_header_index(normalized, ("azimuth", "az"))
    elevation_index = _first_header_index(normalized, ("elevation", "elev", "el"))
    angle_index = _first_header_index(normalized, ("theta",))
    if angle_index is None:
        angle_index = _first_header_index(normalized, ("angle",))
    if angle_index is None:
        angle_index = _first_non_metadata_deg_header_index(header)
    if angle_index is None and azimuth_index is None:
        raise ValueError("DBF dictionary file must contain an angle column.")
    return angle_index if angle_index is not None else azimuth_index, azimuth_index, elevation_index


def _first_non_metadata_deg_header_index(header: list[str]) -> int | None:
    for index, column in enumerate(header):
        normalized = _normalize_header(column)
        if "deg" in normalized and not _looks_like_metadata_column(column):
            return index
    return None


def _dbf_data_column_indices(
    header: list[str],
    angle_indices: set[int],
    angle_index: int | None,
) -> list[int]:
    if angle_index is not None:
        after_angle = [
            index
            for index in range(angle_index + 1, len(header))
            if index not in angle_indices and not _looks_like_metadata_column(header[index])
        ]
        if after_angle:
            return after_angle
    return [
        index
        for index, column in enumerate(header)
        if index not in angle_indices and not _looks_like_metadata_column(column)
    ]


def _first_header_index(headers: list[str], keywords: tuple[str, ...]) -> int | None:
    for index, header in enumerate(headers):
        tokens = set(header.split())
        if any(keyword in tokens or keyword in header for keyword in keywords):
            return index
    return None


def _infer_channel_mode(
    array: AntennaArray,
    column_names: tuple[str, ...],
    column_count: int,
    preferred_mode: str | None = None,
) -> str:
    if preferred_mode in {CHANNEL_MODE_PHYSICAL, CHANNEL_MODE_VIRTUAL}:
        _validate_channel_mode_column_count(array, column_count, preferred_mode)
        return preferred_mode

    if _has_physical_channel_headers(array, column_names):
        return CHANNEL_MODE_PHYSICAL
    if _virtual_column_order(array, column_names) is not None:
        return CHANNEL_MODE_VIRTUAL

    physical_count = len(array.tx) + len(array.rx)
    virtual_count = len(array.tx) * len(array.rx)
    if column_count == physical_count and column_count != virtual_count:
        return CHANNEL_MODE_PHYSICAL
    if column_count == virtual_count and column_count != physical_count:
        return CHANNEL_MODE_VIRTUAL
    if column_count == physical_count == virtual_count:
        return CHANNEL_MODE_PHYSICAL
    raise ValueError(
        f"DBF dictionary has {column_count} channel columns; expected "
        f"{virtual_count} virtual columns or "
        f"{physical_count} physical Tx/Rx columns."
    )


def _validate_channel_mode_column_count(
    array: AntennaArray,
    column_count: int,
    channel_mode: str,
) -> None:
    tx_count = len(array.tx)
    rx_count = len(array.rx)
    physical_counts = {tx_count + rx_count}
    if tx_count == 1:
        physical_counts.add(rx_count)
    virtual_count = tx_count * rx_count
    expected_counts = (
        {virtual_count}
        if channel_mode == CHANNEL_MODE_VIRTUAL
        else physical_counts
    )
    if column_count in expected_counts:
        return
    expected_text = " or ".join(str(count) for count in sorted(expected_counts))
    label = "virtual" if channel_mode == CHANNEL_MODE_VIRTUAL else "physical"
    raise ValueError(
        f"DBF dictionary has {column_count} channel columns; expected "
        f"{expected_text} {label} channel columns."
    )


def _split_physical_columns(
    array: AntennaArray,
    rows: np.ndarray,
    column_names: tuple[str, ...],
) -> tuple[np.ndarray, np.ndarray]:
    name_to_index = {_normalize_header(name): index for index, name in enumerate(column_names)}
    tx_indices = [name_to_index.get(_normalize_header(point.name)) for point in array.tx]
    rx_indices = [name_to_index.get(_normalize_header(point.name)) for point in array.rx]
    if all(index is not None for index in tx_indices + rx_indices):
        return rows[:, tx_indices], rows[:, rx_indices]
    tx_count = len(array.tx)
    rx_count = len(array.rx)
    if tx_count == 1 and rows.shape[1] == rx_count:
        tx_rows = np.ones((rows.shape[0], 1), dtype=complex)
        return tx_rows, rows[:, :rx_count]
    return rows[:, :tx_count], rows[:, tx_count : tx_count + rx_count]


def _reorder_virtual_columns(
    array: AntennaArray,
    rows: np.ndarray,
    column_names: tuple[str, ...],
) -> np.ndarray:
    indices = _virtual_column_order(array, column_names)
    if indices is None:
        return rows
    return rows[:, indices]


def _has_physical_channel_headers(
    array: AntennaArray,
    column_names: tuple[str, ...],
) -> bool:
    normalized = {_normalize_header(name) for name in column_names}
    physical_names = {_normalize_header(point.name) for point in (*array.tx, *array.rx)}
    return physical_names.issubset(normalized)


def _virtual_column_order(
    array: AntennaArray,
    column_names: tuple[str, ...],
) -> list[int] | None:
    name_to_index = {_normalize_header(name): index for index, name in enumerate(column_names)}
    indices = []
    for virtual_index, point in enumerate(array.virtual_points()):
        candidates = (
            f"{point.tx_name}_{point.rx_name}",
            f"{point.tx_name}-{point.rx_name}",
            f"{point.tx_name}{point.rx_name}",
            f"v{virtual_index + 1}",
            f"virtual{virtual_index + 1}",
        )
        index = next(
            (name_to_index[_normalize_header(name)] for name in candidates if _normalize_header(name) in name_to_index),
            None,
        )
        if index is None:
            return None
        indices.append(index)
    return indices


def _parse_complex_or_float(value: object) -> complex:
    text = str(value).strip()
    if not text:
        raise ValueError("blank value")
    normalized = (
        text.replace("i", "j")
        .replace("I", "j")
        .replace("−", "-")
        .replace("，", ",")
        .strip()
    )
    normalized = normalized.replace(" ", "")
    if "∠" in normalized:
        magnitude_text, phase_text = normalized.split("∠", 1)
        return float(magnitude_text) * np.exp(1j * np.radians(float(phase_text)))
    try:
        return complex(normalized)
    except ValueError:
        return complex(_parse_float(text), 0.0)


def _parse_float(value: object) -> float:
    text = str(value).strip().replace(",", ".")
    if not text:
        raise ValueError("blank value")
    return float(text)


def _looks_complex(value: object) -> bool:
    text = str(value).strip().lower()
    return "j" in text or "i" in text or "∠" in text


def _looks_like_metadata_column(value: str) -> bool:
    header = _normalize_header(value)
    tokens = set(header.split())
    return bool(tokens & {"freq", "frequency", "phi"}) or "freq" in header


def _normalize_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.strip().lower()).strip()
